
import pandas as pd
import googlemaps
import re
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import numpy as np

# Chargez votre tableau Excel dans une dataframe pandas
df = pd.read_excel("/Users/quentinschneiter/Desktop/Code/Mobility Plan API/Tableau/Input_1.xlsx", sheet_name="Liste employés")

# Chargez les données du tableau EcoBilan Mobilité
ecobilan = pd.read_excel("/Users/quentinschneiter/Desktop/Code/Mobility Plan API/Tableau/EcoBilan Mobilité.xlsx", sheet_name="Données Plan mobilité")

# Récupérez les facteurs d'émission en kg CO2 équivalent par km
emission_factors = {
    "voiture_essence": ecobilan.loc[63, "Total kg CO2-eq"],
    "voiture_diesel": ecobilan.loc[65, "Total kg CO2-eq"],
    "voiture_electrique": ecobilan.loc[66, "Total kg CO2-eq"],
    "TP": (ecobilan.loc[47, "Total kg CO2-eq"] * 3/4 + ecobilan.loc[31, "Total kg CO2-eq"] * 1/4),
    "VAE_25km/h": 0.005032,
    "Speedelec_45km/h": 0.008288,
    "Motocycle thermique": ecobilan.loc[74, "Total kg CO2-eq"],
    "Motocycle électrique": ecobilan.loc[75, "Total kg CO2-eq"]
    
}

# Créez une instance de la classe Client Google Maps
gmaps = googlemaps.Client(key="AIzaSyBWx7_Me1G-PShLCY4bUtBuOq0yDotvoxc")

def get_travel_time(origin, destination, mode):
    if pd.isnull(origin):
        return "Adresse manquante"
    try:
        origin = origin + ", Suisse"
        destination = destination + ", Suisse"
        result = gmaps.directions(origin, destination, mode=mode)
        if not result:
            return "Adresse non trouvée"
        duration_text = result[0]["legs"][0]["duration"]["text"]
        hours = 0
        minutes = 0
        match = re.search(r'(\d+) hour', duration_text)
        if match:
            hours = int(match.group(1))
        match = re.search(r'(\d+) min', duration_text)
        if match:
            minutes = int(match.group(1))
        return "{:02d}:{:02d}".format(hours, minutes)
    except Exception as e:
        return f"Erreur : {str(e)}"

# Calculez les temps de trajet pour les différents modes de transport
df["Temps de trajet en voiture"] = df.apply(lambda row: get_travel_time(row["Adresse"], "Chemin des Esserts 5, 1024 Ecublens", "driving"), axis=1)
df["Temps de trajet à pied"] = df.apply(lambda row: get_travel_time(row["Adresse"], "Chemin des Esserts 5, 1024 Ecublens", "walking"), axis=1)
df["Temps de trajet en transports publics"] = df.apply(lambda row: get_travel_time(row["Adresse"], "Chemin des Esserts 5, 1024 Ecublens", "transit"), axis=1)
df["Temps de trajet en vélo"] = df.apply(lambda row: get_travel_time(row["Adresse"], "Chemin des Esserts 5, 1024 Ecublens", "bicycling"), axis=1)
df["Temps de trajet en motocycle"] = df.apply(lambda row: get_travel_time(row["Adresse"], "Chemin des Esserts 5, 1024 Ecublens", "driving"), axis=1)

# Convertissez vos temps de trajet en minutes
for column in ["Temps de trajet en voiture", "Temps de trajet en transports publics", "Temps de trajet en vélo", "Temps de trajet à pied", "Temps de trajet en motocycle"]:
    df[column] = df[column].apply(lambda x: int(x.split(':')[0]) * 60 + int(x.split(':')[1]) if isinstance(x, str) else x)

# Calcul du temps de trajet en VAE 25 km/h et en Speedelec 45 km/h
df["Temps de trajet en VAE 25 km/h"] = (16/22) * df["Temps de trajet en vélo"]
df["Temps de trajet en Speedelec 45 km/h"] = (16/32) * df["Temps de trajet en vélo"]

# Assurez-vous que les valeurs sont arrondies au nombre entier le plus proche
df["Temps de trajet en VAE 25 km/h"] = df["Temps de trajet en VAE 25 km/h"].round()
df["Temps de trajet en Speedelec 45 km/h"] = df["Temps de trajet en Speedelec 45 km/h"].round()

# Ajout des colonnes de différence de temps en fonction du véhicule principal
def calculate_difference(row):
    vehicle = row["Véhicule principal"]
    vehicle_time_columns = {
        "Voiture à essence": "Temps de trajet en voiture",
        "Voiture au diesel": "Temps de trajet en voiture",
        "Voiture électrique": "Temps de trajet en voiture",
        "Pieds": "Temps de trajet à pied",
        "Transports publics": "Temps de trajet en transports publics",
        "Vélo": "Temps de trajet en vélo",
        "VAE 25 km/h": "Temps de trajet en VAE 25 km/h",
        "Speedelec 45 km/h": "Temps de trajet en Speedelec 45 km/h",
        "Motocycle thermique":"Temps de trajet en motocycle",
        "Motocycle électrique":"Temps de trajet en motocycle"
    }
    if pd.isnull(vehicle):
        vehicle = "Voiture à essence"
    return row[vehicle_time_columns[vehicle]]

df["Différence temps de trajet en voiture"] = df["Temps de trajet en voiture"] - df.apply(calculate_difference, axis=1)
df["Différence temps de trajet à pied"] = df["Temps de trajet à pied"] - df.apply(calculate_difference, axis=1)
df["Différence temps de trajet en transports publics"] = df["Temps de trajet en transports publics"] - df.apply(calculate_difference, axis=1)
df["Différence temps de trajet en vélo"] = df["Temps de trajet en vélo"] - df.apply(calculate_difference, axis=1)
df["Différence temps de trajet en VAE 25 km/h"] = df["Temps de trajet en VAE 25 km/h"] - df.apply(calculate_difference, axis=1)
df["Différence temps de trajet en Speedelec 45 km/h"] = df["Temps de trajet en Speedelec 45 km/h"] - df.apply(calculate_difference, axis=1)
df["Différence temps de trajet en motocycle"] = df["Temps de trajet en motocycle"] - df.apply(calculate_difference, axis=1)

def get_travel_distance(origin, destination, mode):
    if pd.isnull(origin):
        return "Adresse manquante"
    try:
        origin = origin + ", Suisse"
        destination = destination + ", Suisse"
        result = gmaps.directions(origin, destination, mode=mode)
        if not result:
            return "Adresse non trouvée"
        distance_text = result[0]["legs"][0]["distance"]["text"]
        distance = float(re.search(r'(\d+\.?\d*)', distance_text).group(1))
        if ' km' in distance_text:
            return distance
        elif ' m' in distance_text:
            distance /= 1000.0  # Conversion de mètres en kilomètres
            return distance
    except Exception as e:
        return f"Erreur : {str(e)}"

# Ajout des colonnes de distance pour les différents modes de transport
modes = ["driving", "walking", "transit", "bicycling", "driving"]
for mode, col_name in zip(modes, ["Distance en voiture", "Distance à pied", "Distance en transports publics", "Distance en vélo", "Distance en motocycle"]):
    df[col_name] = df.apply(lambda row: get_travel_distance(row["Adresse"], "Chemin des Esserts 5, 1024 Ecublens", mode), axis=1)


# Ajout des colonnes d'émissions de GES pour les différents modes de transport
df["kg CO2-eq. en voiture à essence"] = df["Distance en voiture"] * emission_factors["voiture_essence"]
df["kg CO2-eq. en voiture au diesel"] = df["Distance en voiture"] * emission_factors["voiture_diesel"]
df["kg CO2-eq. en voiture électrique"] = df["Distance en voiture"] * emission_factors["voiture_electrique"]
df["kg CO2-eq. à pied"] = df["Distance à pied"] * 0
df["kg CO2-eq. en TP"] = df["Distance en transports publics"] * emission_factors["TP"]
df["kg CO2-eq. en vélo"] = df["Distance en vélo"] * 0
df["kg CO2-eq. en VAE 25km/h"] = df["Distance en vélo"] * emission_factors["VAE_25km/h"]
df["kg CO2-eq. en Speedelec 45km/h"] = df["Distance en vélo"] * emission_factors["Speedelec_45km/h"]
df["kg CO2-eq. en motocycle thermique"] = df["Distance en motocycle"] * emission_factors["Motocycle thermique"]
df["kg CO2-eq. en motocycle électrique"] = df["Distance en motocycle"] * emission_factors["Motocycle électrique"]

def calculate_co2_difference(row):
    vehicle = row["Véhicule principal"]
    vehicle_co2_columns = {
        "Voiture à essence": "kg CO2-eq. en voiture à essence",
        "Voiture au diesel": "kg CO2-eq. en voiture au diesel",
        "Voiture électrique": "kg CO2-eq. en voiture électrique",
        "Pieds": "kg CO2-eq. à pied",
        "Transports publics": "kg CO2-eq. en TP",
        "Vélo": "kg CO2-eq. en vélo",
        "VAE 25 km/h": "kg CO2-eq. en VAE 25km/h",
        "Speedelec 45 km/h": "kg CO2-eq. en Speedelec 45km/h",
        "Motocycle thermique": "kg CO2-eq. en motocycle thermique",
        "Motocycle électrique": "kg CO2-eq. en motocycle électrique"
    }
    if pd.isnull(vehicle):
        vehicle = "Voiture à essence"
    return row[vehicle_co2_columns[vehicle]]

df["Différence kg CO2-eq. en voiture à essence"] = df["kg CO2-eq. en voiture à essence"] - df.apply(calculate_co2_difference, axis=1)
df["Différence kg CO2-eq. en voiture au diesel"] = df["kg CO2-eq. en voiture au diesel"] - df.apply(calculate_co2_difference, axis=1)
df["Différence kg CO2-eq. en voiture électrique"] = df["kg CO2-eq. en voiture électrique"] - df.apply(calculate_co2_difference, axis=1)
df["Différence kg CO2-eq. à pied"] = df["kg CO2-eq. à pied"] - df.apply(calculate_co2_difference, axis=1)
df["Différence kg CO2-eq. en TP"] = df["kg CO2-eq. en TP"] - df.apply(calculate_co2_difference, axis=1)
df["Différence kg CO2-eq. en vélo"] = df["kg CO2-eq. en vélo"] - df.apply(calculate_co2_difference, axis=1)
df["Différence kg CO2-eq. en VAE 25km/h"] = df["kg CO2-eq. en VAE 25km/h"] - df.apply(calculate_co2_difference, axis=1)
df["Différence kg CO2-eq. en Speedelec 45km/h"] = df["kg CO2-eq. en Speedelec 45km/h"] - df.apply(calculate_co2_difference, axis=1)
df["Différence kg CO2-eq. en motocycle thermique"] = df["kg CO2-eq. en motocycle thermique"] - df.apply(calculate_co2_difference, axis=1)
df["Différence kg CO2-eq. en motocycle électrique"] = df["kg CO2-eq. en motocycle électrique"] - df.apply(calculate_co2_difference, axis=1)

# Créez une colonne pour le facteur d'ajustement
df["Facteur d'ajustement"] = 502 * ((df["Pourcentage"] - (df["Nbre jours télétravail"] * 20)) / 100)

# Multipliez les colonnes d'émissions de CO2 et de différence d'émissions de CO2 par le facteur d'ajustement
cols_emissions = [
    "kg CO2-eq. en voiture à essence",
    "kg CO2-eq. en voiture au diesel",
    "kg CO2-eq. en voiture électrique",
    "kg CO2-eq. à pied",
    "kg CO2-eq. en TP",
    "kg CO2-eq. en vélo",
    "kg CO2-eq. en VAE 25km/h",
    "kg CO2-eq. en Speedelec 45km/h",
    "Différence kg CO2-eq. en voiture à essence",
    "Différence kg CO2-eq. en voiture au diesel",
    "Différence kg CO2-eq. en voiture électrique",
    "Différence kg CO2-eq. à pied",
    "Différence kg CO2-eq. en TP",
    "Différence kg CO2-eq. en vélo",
    "Différence kg CO2-eq. en VAE 25km/h",
    "Différence kg CO2-eq. en Speedelec 45km/h",
    "Différence kg CO2-eq. en motocycle thermique",
    "Différence kg CO2-eq. en motocycle électrique"
]

for col in cols_emissions:
    df[col + " annuelle"] = df[col] * df["Facteur d'ajustement"]

# Ajout des fonctions pour déterminer si la différence de temps de trajet est colorée en vert et pour calculer la réduction potentielle de CO2-eq.
def is_green(row, col_name):
    ratio = row[col_name] / row["Temps de trajet en voiture"]
    return row[col_name] <= 10 or (ratio <= 4/3 and row[col_name] <= 20)

# Dictionnaire pour mapper les noms de colonnes de différence de temps de trajet aux noms de mode de transport
diff_time_to_mode = {
    "Différence temps de trajet en voiture": "Voiture",
    "Différence temps de trajet à pied": "Pieds",
    "Différence temps de trajet en transports publics": "Transports publics",
    "Différence temps de trajet en vélo": "Vélo",
    "Différence temps de trajet en VAE 25 km/h": "VAE 25 km/h",
    "Différence temps de trajet en Speedelec 45 km/h": "Speedelec 45 km/h",
    "Différence temps de trajet en motocycle": "Motocycle",
}

def get_mode_with_low_constraints(row):
    vehicle_diff_co2_columns = {
        "Différence kg CO2-eq. en voiture à essence annuelle": "Différence temps de trajet en voiture",
        "Différence kg CO2-eq. en voiture au diesel annuelle": "Différence temps de trajet en voiture",
        "Différence kg CO2-eq. en voiture électrique annuelle": "Différence temps de trajet en voiture",
        "Différence kg CO2-eq. à pied annuelle": "Différence temps de trajet à pied",
        "Différence kg CO2-eq. en TP annuelle": "Différence temps de trajet en transports publics",
        "Différence kg CO2-eq. en vélo annuelle": "Différence temps de trajet en vélo",
        "Différence kg CO2-eq. en VAE 25km/h annuelle": "Différence temps de trajet en VAE 25 km/h",
        "Différence kg CO2-eq. en Speedelec 45km/h annuelle": "Différence temps de trajet en Speedelec 45 km/h",
        "Différence kg CO2-eq. en motocycle thermique": "Différence temps de trajet en motocycle",
        "Différence kg CO2-eq. en motocycle électrique": "Différence temps de trajet en motocycle"
    }
    green_modes = [col for col, diff_col in vehicle_diff_co2_columns.items() if is_green(row, diff_col)]
    if not green_modes:
        return None
    return vehicle_diff_co2_columns[min(green_modes, key=lambda col: row[col])]

def calculate_potential_reduction(row):
    vehicle_diff_co2_columns = {
        "Différence kg CO2-eq. en voiture à essence annuelle": "Différence temps de trajet en voiture",
        "Différence kg CO2-eq. en voiture au diesel annuelle": "Différence temps de trajet en voiture",
        "Différence kg CO2-eq. en voiture électrique annuelle": "Différence temps de trajet en voiture",
        "Différence kg CO2-eq. à pied annuelle": "Différence temps de trajet à pied",
        "Différence kg CO2-eq. en TP annuelle": "Différence temps de trajet en transports publics",
        "Différence kg CO2-eq. en vélo annuelle": "Différence temps de trajet en vélo",
        "Différence kg CO2-eq. en VAE 25km/h annuelle": "Différence temps de trajet en VAE 25 km/h",
        "Différence kg CO2-eq. en Speedelec 45km/h annuelle": "Différence temps de trajet en Speedelec 45 km/h",
        "Différence kg CO2-eq. en motocycle thermique": "Différence temps de trajet en motocycle",
        "Différence kg CO2-eq. en motocycle électrique": "Différence temps de trajet en motocycle"
    }
    green_modes = [col for col, diff_col in vehicle_diff_co2_columns.items() if is_green(row, diff_col)]
    if not green_modes:
        return 0
    return min(row[col] for col in green_modes)

# Ajout des colonnes mode de transport et réduction potentielle kg CO2-eq. avec contraintes faibles
df["Mode de transport retenu avec contraintes faibles"] = df.apply(get_mode_with_low_constraints, axis=1)
df["Réduction potentielle kg CO2-eq. avec contraintes faibles"] = df.apply(calculate_potential_reduction, axis=1)

# Mappage pour un affichage plus lisible
mode_display_mapping = {
    "Différence temps de trajet en voiture": "en voiture",
    "Différence temps de trajet à pied": "à pieds",
    "Différence temps de trajet en transports publics": "en transports publics",
    "Différence temps de trajet en vélo": "en vélo",
    "Différence temps de trajet en VAE 25 km/h": "en VAE 25 km/h",
    "Différence temps de trajet en Speedelec 45 km/h": "en Speedelec 45 km/h",
    "Différence temps de trajet en motocycle": "Motocycle"
}


df["Mode de transport retenu avec contraintes faibles"] = df["Mode de transport retenu avec contraintes faibles"].map(mode_display_mapping).fillna(df["Mode de transport retenu avec contraintes faibles"])

# Ajout des fonctions pour déterminer si la différence de temps de trajet est colorée en vert ou en orange et pour calculer la réduction potentielle de CO2-eq.

def is_green_or_orange(row, col_name):
    ratio = row[col_name] / row["Temps de trajet en voiture"]
    return row[col_name] <= 10 or (ratio <= 4/3 and row[col_name] <= 20) or (row[col_name] <= 15 or (ratio <= 3/2 and row[col_name] <= 30))

def get_mode_with_moderate_constraints(row):
    vehicle_diff_co2_columns = {
        "Différence kg CO2-eq. en voiture à essence annuelle": "Différence temps de trajet en voiture",
        "Différence kg CO2-eq. en voiture au diesel annuelle": "Différence temps de trajet en voiture",
        "Différence kg CO2-eq. en voiture électrique annuelle": "Différence temps de trajet en voiture",
        "Différence kg CO2-eq. à pied annuelle": "Différence temps de trajet à pied",
        "Différence kg CO2-eq. en TP annuelle": "Différence temps de trajet en transports publics",
        "Différence kg CO2-eq. en vélo annuelle": "Différence temps de trajet en vélo",
        "Différence kg CO2-eq. en VAE 25km/h annuelle": "Différence temps de trajet en VAE 25 km/h",
        "Différence kg CO2-eq. en Speedelec 45km/h annuelle": "Différence temps de trajet en Speedelec 45 km/h",
        "Différence kg CO2-eq. en motocycle thermique": "Différence temps de trajet en motocycle",
        "Différence kg CO2-eq. en motocycle électrique": "Différence temps de trajet en motocycle"
    }
    green_or_orange_modes = [col for col, diff_col in vehicle_diff_co2_columns.items() if is_green_or_orange(row, diff_col)]
    if not green_or_orange_modes:
        return None
    return vehicle_diff_co2_columns[min(green_or_orange_modes, key=lambda col: row[col])]

def calculate_potential_reduction_green_or_orange(row):
    vehicle_diff_co2_columns = {
        "Différence kg CO2-eq. en voiture à essence annuelle": "Différence temps de trajet en voiture",
        "Différence kg CO2-eq. en voiture au diesel annuelle": "Différence temps de trajet en voiture",
        "Différence kg CO2-eq. en voiture électrique annuelle": "Différence temps de trajet en voiture",
        "Différence kg CO2-eq. à pied annuelle": "Différence temps de trajet à pied",
        "Différence kg CO2-eq. en TP annuelle": "Différence temps de trajet en transports publics",
        "Différence kg CO2-eq. en vélo annuelle": "Différence temps de trajet en vélo",
        "Différence kg CO2-eq. en VAE 25km/h annuelle": "Différence temps de trajet en VAE 25 km/h",
        "Différence kg CO2-eq. en Speedelec 45km/h annuelle": "Différence temps de trajet en Speedelec 45 km/h",
        "Différence kg CO2-eq. en motocycle thermique": "Différence temps de trajet en motocycle",
        "Différence kg CO2-eq. en motocycle électrique": "Différence temps de trajet en motocycle"
    }
    green_or_orange_modes = [col for col, diff_col in vehicle_diff_co2_columns.items() if is_green_or_orange(row, diff_col)]
    if not green_or_orange_modes:
        return 0
    return min(row[col] for col in green_or_orange_modes)

# Ajout des colonnes mode de transport et réduction potentielle kg CO2-eq. avec contraintes modérées

df["Mode de transport retenu avec contraintes modérées"] = df.apply(get_mode_with_moderate_constraints, axis=1)
df["Réduction potentielle kg CO2-eq. avec contraintes modérées"] = df.apply(calculate_potential_reduction_green_or_orange, axis=1)

# Mappage pour un affichage plus lisible
mode_display_mapping = {
    "Différence temps de trajet en voiture": "en voiture",
    "Différence temps de trajet à pied": "à pieds",
    "Différence temps de trajet en transports publics": "en transports publics",
    "Différence temps de trajet en vélo": "en vélo",
    "Différence temps de trajet en VAE 25 km/h": "en VAE 25 km/h",
    "Différence temps de trajet en Speedelec 45 km/h": "en Speedelec 45 km/h",
    "Différence temps de trajet en motocycle": "Motocycle"
}

df["Mode de transport retenu avec contraintes modérées"] = df["Mode de transport retenu avec contraintes modérées"].map(mode_display_mapping).fillna(df["Mode de transport retenu avec contraintes modérées"])

# Enregistrement des résultats dans un fichier Excel
output_path = "/Users/quentinschneiter/Desktop/Code/Mobility Plan API/Tableau/Output_1.xlsx"
df.to_excel(output_path, index=False)

# Mise en forme conditionnelle
wb = load_workbook(output_path)
ws = wb.active

# Dictionnaire pour mapper le véhicule principal aux indices des colonnes correspondantes
vehicle_to_columns = {
    "Voiture à essence": [9, 23, 27, 37, 49, 57],
    "Voiture au diesel": [9, 23, 28, 38, 50, 58],
    "Voiture électrique": [9, 23, 29, 39, 51, 59],
    "Pieds": [10, 24, 30, 40, 52, 60],
    "Transports publics": [11, 25, 31, 41, 53, 61],
    "Vélo": [12, 26, 32, 42, 54, 62],
    "VAE 25 km/h": [13, 26, 33, 43, 55, 63],
    "Speedelec 45 km/h": [14, 26, 34, 44, 56, 64],
    "Motocycle thermique": [15, 27, 35, 45, 57, 65],
    "Motocycle électrique": [15, 27, 36, 46, 58, 66]
    
}

for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
    # Mise en forme conditionnelle basée sur le ratio
    for col_idx in [16, 17, 18, 19, 20, 21, 22]:
        cell = ws.cell(row=row_idx, column=col_idx)
        if cell.value is not None:
            ratio = cell.value / df.loc[row_idx - 2, "Temps de trajet en voiture"]
            if np.isnan(ratio):
                fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
            elif cell.value <= 10 or (ratio <= 4/3 and cell.value <= 20):
                fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
            elif cell.value <= 15 or (ratio <= 3/2 and cell.value <= 30):
                fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
            else:
                fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
            cell.fill = fill

    # Mise en forme conditionnelle basée sur le véhicule principal
    primary_vehicle = ws.cell(row=row_idx, column=8).value
    if primary_vehicle in vehicle_to_columns:
        for col_idx in vehicle_to_columns[primary_vehicle]:
            cell = ws.cell(row=row_idx, column=col_idx)
            fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            cell.fill = fill

# Calculer le total pour la colonne "Réduction potentielle kg CO2-eq. avec contraintes faibles"
total_low_constraints = sum([cell.value for cell in ws['BP'] if isinstance(cell.value, (int, float))])

# Calculer le total pour la colonne "Réduction potentielle kg CO2-eq. avec contraintes modérées"
total_moderate_constraints = sum([cell.value for cell in ws['BR'] if isinstance(cell.value, (int, float))])

# Écrire les totaux au bas des colonnes respectives
ws.cell(row=ws.max_row + 1, column=68, value=total_low_constraints)
ws.cell(row=ws.max_row, column=70, value=total_moderate_constraints)

wb.save(output_path)











